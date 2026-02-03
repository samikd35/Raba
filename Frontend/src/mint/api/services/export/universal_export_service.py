"""
Universal Export Service

This service provides comprehensive export functionality for all admin dashboard data,
including PDF generation with charts and branding, CSV exports, and scheduled reports.
"""

import logging
import io
import csv
import json
import uuid
from typing import Optional, List, Dict, Any, Union
from datetime import datetime, timedelta, date
from pathlib import Path
import base64

# PDF generation imports
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab not available - PDF export will be disabled")

# Chart generation imports
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    logging.warning("Matplotlib not available - chart generation will be disabled")

# Excel export imports
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("OpenPyXL not available - Excel export will be disabled")

from ..communication.email_service import email_service
# Import services - these would be properly initialized in production
# For now, we'll create mock services to avoid import issues
try:
    from ..user.analytics import user_analytics_service
except ImportError:
    print("Using mock user_analytics_service for testing")
    from unittest.mock import MagicMock
    user_analytics_service = MagicMock()

try:
    from .system_metrics_service import system_metrics_service
except ImportError:
    print("Using mock system_metrics_service for testing")
    from unittest.mock import MagicMock
    system_metrics_service = MagicMock()

try:
    from .audit_service import audit_service
except ImportError:
    print("Using mock audit_service for testing")
    from unittest.mock import MagicMock
    audit_service = MagicMock()

try:
    from .user_engagement_service import engagement_service
except ImportError:
    print("Using mock engagement_service for testing")
    try:
        from .mock_services import mock_engagement_service
        engagement_service = mock_engagement_service
    except ImportError:
        from unittest.mock import MagicMock
        engagement_service = MagicMock()
from ....schemas.schemas import (
    DataExportRequest,
    DataExportResult,
    ExportFormat,
    ScheduledReportConfig,
    ScheduledReport
)
# Type aliases for ReportLab components
from typing import Any
ReportLabImage = Any

logger = logging.getLogger(__name__)


class UniversalExportService:
    """Service for exporting all types of admin dashboard data."""
    
    def __init__(self):
        self.email_service = email_service
        self.user_analytics_service = user_analytics_service
        self.system_metrics_service = system_metrics_service
        self.audit_service = audit_service
        self.engagement_service = engagement_service
        
        # Data type handlers
        self.data_handlers = {
            'users': self._fetch_user_data,
            'user_analytics': self._fetch_user_analytics_data,
            'system_metrics': self._fetch_system_metrics_data,
            'agent_status': self._fetch_agent_status_data,
            'api_health': self._fetch_api_health_data,
            'audit_logs': self._fetch_audit_logs_data,
            'engagement': self._fetch_engagement_data,
            'business_intelligence': self._fetch_business_intelligence_data,
            'session_management': self._fetch_session_data
        }
        
        # Format handlers
        self.format_handlers = {
            ExportFormat.CSV: self._export_csv,
            ExportFormat.JSON: self._export_json,
            ExportFormat.PDF: self._export_pdf,
            ExportFormat.EXCEL: self._export_excel
        }
        
    async def export_data(
        self,
        export_request: DataExportRequest
    ) -> DataExportResult:
        """
        Export admin dashboard data in the specified format.
        
        Args:
            export_request: Export configuration and parameters
            
        Returns:
            Export result with file data or download URL
        """
        try:
            export_id = str(uuid.uuid4())
            logger.info(f"Starting data export {export_id}: {export_request.data_type} as {export_request.format.value}")
            
            # Validate data type
            if export_request.data_type not in self.data_handlers:
                raise ValueError(f"Unsupported data type: {export_request.data_type}")
            
            # Validate format
            if export_request.format not in self.format_handlers:
                raise ValueError(f"Unsupported export format: {export_request.format}")
            
            # Fetch the data
            data = await self.data_handlers[export_request.data_type](export_request)
            
            # Generate export based on format
            result = await self.format_handlers[export_request.format](data, export_request, export_id)
            
            # Send email if requested
            if export_request.email_delivery and export_request.email_recipients:
                email_sent = await self._send_export_email(result, export_request)
                result.email_sent = email_sent
            
            logger.info(f"Successfully exported data {export_id}: {export_request.data_type} as {export_request.format.value}")
            return result
            
        except Exception as e:
            logger.error(f"Error exporting data: {str(e)}")
            raise
    
    # Data fetching methods
    
    async def _fetch_user_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch user management data."""
        try:
            if not self.user_analytics_service:
                # Return mock data for testing
                return {
                    'data_type': 'users',
                    'users': [
                        {
                            'user_id': 'user_1',
                            'email': 'user1@example.com',
                            'display_name': 'Test User 1',
                            'created_at': '2024-01-01T00:00:00Z',
                            'last_login': '2024-01-20T12:00:00Z',
                            'login_count': 25,
                            'total_requests': 150,
                            'total_reports_generated': 45,
                            'total_screen_time_minutes': 1200,
                            'subscription_status': 'premium',
                            'geographic_location': 'US',
                            'engagement_score': 85.5
                        },
                        {
                            'user_id': 'user_2',
                            'email': 'user2@example.com',
                            'display_name': 'Test User 2',
                            'created_at': '2024-01-15T00:00:00Z',
                            'last_login': '2024-01-21T09:30:00Z',
                            'login_count': 12,
                            'total_requests': 75,
                            'total_reports_generated': 20,
                            'total_screen_time_minutes': 600,
                            'subscription_status': 'free',
                            'geographic_location': 'UK',
                            'engagement_score': 72.3
                        }
                    ],
                    'total_count': 2,
                    'export_metadata': {
                        'generated_at': datetime.utcnow().isoformat(),
                        'filters_applied': export_request.filters,
                        'note': 'Mock data for testing'
                    }
                }
            
            # Get user analytics summary
            users = await self.user_analytics_service.get_user_analytics(
                limit=10000,  # Large limit for export
                offset=0
            )
            
            return {
                'data_type': 'users',
                'users': [user.dict() for user in users.users],
                'total_count': users.total_count,
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat(),
                    'filters_applied': export_request.filters
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching user data: {str(e)}")
            raise
    
    async def _fetch_user_analytics_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch detailed user analytics data."""
        try:
            if not self.user_analytics_service:
                # Return mock analytics data
                return {
                    'data_type': 'user_analytics',
                    'analytics': [
                        {
                            'time_bucket': '2024-01-20',
                            'new_users': 5,
                            'active_users': 25,
                            'total_logins': 45,
                            'total_requests': 120,
                            'total_reports': 35,
                            'avg_engagement_score': 78.5
                        },
                        {
                            'time_bucket': '2024-01-21',
                            'new_users': 3,
                            'active_users': 28,
                            'total_logins': 52,
                            'total_requests': 135,
                            'total_reports': 42,
                            'avg_engagement_score': 81.2
                        }
                    ],
                    'export_metadata': {
                        'generated_at': datetime.utcnow().isoformat(),
                        'date_range': {
                            'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                            'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                        },
                        'note': 'Mock data for testing'
                    }
                }
            
            # Get aggregated analytics
            analytics = await self.user_analytics_service.get_aggregated_analytics(
                start_date=export_request.start_date,
                end_date=export_request.end_date,
                interval="1 day"
            )
            
            return {
                'data_type': 'user_analytics',
                'analytics': [analytic.dict() for analytic in analytics],
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat(),
                    'date_range': {
                        'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                        'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                    }
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching user analytics data: {str(e)}")
            raise
    
    async def _fetch_system_metrics_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch system metrics data."""
        try:
            if not self.system_metrics_service:
                # Return mock system metrics data
                return {
                    'data_type': 'system_metrics',
                    'metrics': [
                        {
                            'time_bucket': '2024-01-21T10:00:00Z',
                            'metric_type': 'infrastructure',
                            'metric_name': 'cpu_usage',
                            'avg_value': 45.2,
                            'min_value': 32.1,
                            'max_value': 67.8,
                            'count_values': 60
                        },
                        {
                            'time_bucket': '2024-01-21T10:00:00Z',
                            'metric_type': 'infrastructure',
                            'metric_name': 'memory_usage',
                            'avg_value': 68.5,
                            'min_value': 55.2,
                            'max_value': 82.1,
                            'count_values': 60
                        }
                    ],
                    'export_metadata': {
                        'generated_at': datetime.utcnow().isoformat(),
                        'date_range': {
                            'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                            'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                        },
                        'note': 'Mock data for testing'
                    }
                }
            
            # Get aggregated metrics
            metrics = await self.system_metrics_service.get_aggregated_metrics(
                start_date=export_request.start_date,
                end_date=export_request.end_date,
                interval="1 hour"
            )
            
            return {
                'data_type': 'system_metrics',
                'metrics': [metric.dict() for metric in metrics],
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat(),
                    'date_range': {
                        'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                        'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                    }
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching system metrics data: {str(e)}")
            raise
    
    async def _fetch_agent_status_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch agent status data."""
        try:
            # Get all agent statuses
            agents = await self.system_metrics_service.get_all_agent_statuses()
            
            return {
                'data_type': 'agent_status',
                'agents': [agent.dict() for agent in agents],
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching agent status data: {str(e)}")
            raise
    
    async def _fetch_api_health_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch API health data."""
        try:
            # Get API health summaries
            health_data = await self.system_metrics_service.get_api_health_summary()
            
            return {
                'data_type': 'api_health',
                'health_data': [health.dict() for health in health_data],
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching API health data: {str(e)}")
            raise
    
    async def _fetch_audit_logs_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch audit logs data."""
        try:
            # Get audit logs
            logs = await self.audit_service.get_audit_logs(
                limit=10000,  # Large limit for export
                offset=0,
                start_date=export_request.start_date,
                end_date=export_request.end_date
            )
            
            return {
                'data_type': 'audit_logs',
                'logs': [log.dict() for log in logs.logs],
                'total_count': logs.total_count,
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat(),
                    'date_range': {
                        'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                        'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                    }
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching audit logs data: {str(e)}")
            raise
    
    async def _fetch_engagement_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch engagement analytics data."""
        try:
            if not self.engagement_service:
                # Return mock engagement data
                return {
                    'data_type': 'engagement',
                    'overview': {
                        'total_users': 150,
                        'active_users': 85,
                        'new_users': 12,
                        'total_sessions': 320,
                        'avg_session_duration': 15.5,
                        'total_page_views': 1250,
                        'avg_pages_per_session': 3.9,
                        'total_clicks': 890,
                        'total_reports_generated': 145,
                        'total_links_clicked': 234,
                        'avg_engagement_score': 76.8,
                        'bounce_rate': 25.3
                    },
                    'daily_metrics': [
                        {
                            'date': '2024-01-20',
                            'total_users': 45,
                            'active_users': 32,
                            'total_sessions': 78,
                            'total_page_views': 312,
                            'total_clicks': 189,
                            'total_reports_generated': 34,
                            'avg_session_duration': 14.2,
                            'avg_engagement_score': 75.5
                        },
                        {
                            'date': '2024-01-21',
                            'total_users': 52,
                            'active_users': 38,
                            'total_sessions': 89,
                            'total_page_views': 356,
                            'total_clicks': 201,
                            'total_reports_generated': 41,
                            'avg_session_duration': 16.8,
                            'avg_engagement_score': 78.1
                        }
                    ],
                    'export_metadata': {
                        'generated_at': datetime.utcnow().isoformat(),
                        'date_range': {
                            'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                            'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                        },
                        'note': 'Mock data for testing'
                    }
                }
            
            # Get engagement overview
            overview = await self.engagement_service.get_engagement_overview(
                start_date=export_request.start_date,
                end_date=export_request.end_date
            )
            
            # Get daily metrics
            daily_metrics = await self.engagement_service.get_daily_metrics(
                start_date=export_request.start_date,
                end_date=export_request.end_date,
                limit=365
            )
            
            return {
                'data_type': 'engagement',
                'overview': overview.dict(),
                'daily_metrics': [metric.dict() for metric in daily_metrics],
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat(),
                    'date_range': {
                        'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                        'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                    }
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching engagement data: {str(e)}")
            raise
    
    async def _fetch_business_intelligence_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch business intelligence data."""
        try:
            # This would aggregate data from multiple sources
            # For now, return placeholder structure
            return {
                'data_type': 'business_intelligence',
                'usage_trends': [],
                'industry_analytics': [],
                'performance_benchmarks': [],
                'cost_analysis': [],
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat(),
                    'note': 'Business intelligence data aggregation in development'
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching business intelligence data: {str(e)}")
            raise
    
    async def _fetch_session_data(self, export_request: DataExportRequest) -> Dict[str, Any]:
        """Fetch session management data."""
        try:
            # This would fetch active session data
            # For now, return placeholder structure
            return {
                'data_type': 'session_management',
                'active_sessions': [],
                'session_history': [],
                'export_metadata': {
                    'generated_at': datetime.utcnow().isoformat(),
                    'note': 'Session management data in development'
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching session data: {str(e)}")
            raise
    
    # Export format methods
    
    async def _export_csv(
        self,
        data: Dict[str, Any],
        export_request: DataExportRequest,
        export_id: str
    ) -> DataExportResult:
        """Export data as CSV."""
        try:
            output = io.StringIO()
            
            # Create CSV content based on data type
            if export_request.data_type == 'users':
                csv_content = self._create_users_csv(data)
            elif export_request.data_type == 'user_analytics':
                csv_content = self._create_user_analytics_csv(data)
            elif export_request.data_type == 'system_metrics':
                csv_content = self._create_system_metrics_csv(data)
            elif export_request.data_type == 'audit_logs':
                csv_content = self._create_audit_logs_csv(data)
            elif export_request.data_type == 'engagement':
                csv_content = self._create_engagement_csv(data)
            else:
                csv_content = self._create_generic_csv(data)
            
            filename = f"{export_request.data_type}_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return DataExportResult(
                export_id=export_id,
                data_type=export_request.data_type,
                format=ExportFormat.CSV,
                filename=filename,
                content=csv_content.encode('utf-8'),
                content_type='text/csv',
                size=len(csv_content.encode('utf-8')),
                generated_at=datetime.utcnow()
            )
            
        except Exception as e:
            logger.error(f"Error creating CSV export: {str(e)}")
            raise
    
    def _create_users_csv(self, data: Dict[str, Any]) -> str:
        """Create CSV content for users data."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            'User ID', 'Email', 'Display Name', 'Created At', 'Last Login',
            'Login Count', 'Total Requests', 'Total Reports', 'Screen Time (min)',
            'Subscription Status', 'Geographic Location', 'Engagement Score'
        ])
        
        # Data rows
        for user in data.get('users', []):
            writer.writerow([
                user.get('user_id', ''),
                user.get('email', ''),
                user.get('display_name', ''),
                user.get('created_at', ''),
                user.get('last_login', ''),
                user.get('login_count', 0),
                user.get('total_requests', 0),
                user.get('total_reports_generated', 0),
                user.get('total_screen_time_minutes', 0),
                user.get('subscription_status', ''),
                user.get('geographic_location', ''),
                user.get('engagement_score', 0)
            ])
        
        return output.getvalue()
    
    def _create_user_analytics_csv(self, data: Dict[str, Any]) -> str:
        """Create CSV content for user analytics data."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            'Date', 'New Users', 'Active Users', 'Total Logins',
            'Total Requests', 'Total Reports', 'Avg Engagement Score'
        ])
        
        # Data rows
        for analytic in data.get('analytics', []):
            writer.writerow([
                analytic.get('time_bucket', ''),
                analytic.get('new_users', 0),
                analytic.get('active_users', 0),
                analytic.get('total_logins', 0),
                analytic.get('total_requests', 0),
                analytic.get('total_reports', 0),
                analytic.get('avg_engagement_score', 0)
            ])
        
        return output.getvalue()
    
    def _create_system_metrics_csv(self, data: Dict[str, Any]) -> str:
        """Create CSV content for system metrics data."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            'Timestamp', 'Metric Type', 'Metric Name', 'Avg Value',
            'Min Value', 'Max Value', 'Count'
        ])
        
        # Data rows
        for metric in data.get('metrics', []):
            writer.writerow([
                metric.get('time_bucket', ''),
                metric.get('metric_type', ''),
                metric.get('metric_name', ''),
                metric.get('avg_value', 0),
                metric.get('min_value', 0),
                metric.get('max_value', 0),
                metric.get('count_values', 0)
            ])
        
        return output.getvalue()
    
    def _create_audit_logs_csv(self, data: Dict[str, Any]) -> str:
        """Create CSV content for audit logs data."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            'Timestamp', 'Admin User ID', 'Action', 'Target Type',
            'Target ID', 'Success', 'IP Address', 'Details'
        ])
        
        # Data rows
        for log in data.get('logs', []):
            writer.writerow([
                log.get('timestamp', ''),
                log.get('admin_user_id', ''),
                log.get('action', ''),
                log.get('target_type', ''),
                log.get('target_id', ''),
                log.get('success', ''),
                log.get('ip_address', ''),
                json.dumps(log.get('details', {}))
            ])
        
        return output.getvalue()
    
    def _create_engagement_csv(self, data: Dict[str, Any]) -> str:
        """Create CSV content for engagement data."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Overview section
        writer.writerow(['ENGAGEMENT OVERVIEW'])
        overview = data.get('overview', {})
        for key, value in overview.items():
            writer.writerow([key, value])
        
        writer.writerow([])  # Empty row
        
        # Daily metrics section
        writer.writerow(['DAILY METRICS'])
        writer.writerow([
            'Date', 'Total Users', 'Active Users', 'Total Sessions',
            'Total Page Views', 'Total Clicks', 'Reports Generated',
            'Avg Session Duration', 'Engagement Score'
        ])
        
        for metric in data.get('daily_metrics', []):
            writer.writerow([
                metric.get('date', ''),
                metric.get('total_users', 0),
                metric.get('active_users', 0),
                metric.get('total_sessions', 0),
                metric.get('total_page_views', 0),
                metric.get('total_clicks', 0),
                metric.get('total_reports_generated', 0),
                metric.get('avg_session_duration', 0),
                metric.get('avg_engagement_score', 0)
            ])
        
        return output.getvalue()
    
    def _create_generic_csv(self, data: Dict[str, Any]) -> str:
        """Create generic CSV content for any data type."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write metadata
        writer.writerow(['DATA EXPORT'])
        writer.writerow(['Data Type', data.get('data_type', 'unknown')])
        writer.writerow(['Generated At', data.get('export_metadata', {}).get('generated_at', '')])
        writer.writerow([])
        
        # Write raw data as JSON
        writer.writerow(['RAW DATA (JSON)'])
        writer.writerow([json.dumps(data, indent=2, default=str)])
        
        return output.getvalue()
    
    async def _export_json(
        self,
        data: Dict[str, Any],
        export_request: DataExportRequest,
        export_id: str
    ) -> DataExportResult:
        """Export data as JSON."""
        try:
            # Add export metadata
            export_data = {
                'metadata': {
                    'export_id': export_id,
                    'export_date': datetime.utcnow().isoformat(),
                    'data_type': export_request.data_type,
                    'date_range': {
                        'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                        'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                    },
                    'filters': export_request.filters,
                    'export_format': export_request.format.value
                },
                'data': data
            }
            
            json_content = json.dumps(export_data, indent=2, default=str)
            
            filename = f"{export_request.data_type}_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
            
            return DataExportResult(
                export_id=export_id,
                data_type=export_request.data_type,
                format=ExportFormat.JSON,
                filename=filename,
                content=json_content.encode('utf-8'),
                content_type='application/json',
                size=len(json_content.encode('utf-8')),
                generated_at=datetime.utcnow()
            )
            export_data = {
                'export_id': export_id,
                'export_metadata': {
                    'data_type': export_request.data_type,
                    'format': export_request.format.value,
                    'generated_at': datetime.utcnow().isoformat(),
                    'filters': export_request.filters,
                    'date_range': {
                        'start_date': export_request.start_date.isoformat() if export_request.start_date else None,
                        'end_date': export_request.end_date.isoformat() if export_request.end_date else None
                    }
                },
                'data': data
            }
            
            json_content = json.dumps(export_data, indent=2, default=str)
            filename = f"{export_request.data_type}_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
            
            return DataExportResult(
                export_id=export_id,
                data_type=export_request.data_type,
                format=ExportFormat.JSON,
                filename=filename,
                content=json_content.encode('utf-8'),
                content_type='application/json',
                size=len(json_content.encode('utf-8')),
                generated_at=datetime.utcnow()
            )
            
        except Exception as e:
            logger.error(f"Error creating JSON export: {str(e)}")
            raise
    
    async def _export_pdf(
        self,
        data: Dict[str, Any],
        export_request: DataExportRequest,
        export_id: str
    ) -> DataExportResult:
        """Export data as PDF with charts and branding."""
        if not REPORTLAB_AVAILABLE:
            raise Exception("PDF export not available - ReportLab not installed")
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title and branding
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=HexColor('#2c3e50')
            )
            
            data_type_title = export_request.data_type.replace('_', ' ').title()
            story.append(Paragraph(f"MINT Admin Dashboard - {data_type_title} Report", title_style))
            story.append(Spacer(1, 20))
            
            # Report metadata
            metadata_style = styles['Normal']
            story.append(Paragraph(f"<b>Export ID:</b> {export_id}", metadata_style))
            story.append(Paragraph(f"<b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", metadata_style))
            if export_request.start_date and export_request.end_date:
                story.append(Paragraph(f"<b>Date Range:</b> {export_request.start_date} to {export_request.end_date}", metadata_style))
            story.append(Spacer(1, 30))
            
            # Data-specific content
            if export_request.data_type == 'users':
                story.extend(self._create_users_pdf_content(data, styles))
            elif export_request.data_type == 'audit_logs':
                story.extend(self._create_audit_logs_pdf_content(data, styles))
            elif export_request.data_type == 'engagement':
                story.extend(self._create_engagement_pdf_content(data, styles, export_request.include_charts))
            else:
                story.extend(self._create_generic_pdf_content(data, styles))
            
            # Build PDF
            doc.build(story)
            pdf_content = buffer.getvalue()
            buffer.close()
            
            filename = f"{export_request.data_type}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            return DataExportResult(
                export_id=export_id,
                data_type=export_request.data_type,
                format=ExportFormat.PDF,
                filename=filename,
                content=pdf_content,
                content_type='application/pdf',
                size=len(pdf_content),
                generated_at=datetime.utcnow()
            )
            
        except Exception as e:
            logger.error(f"Error creating PDF export: {str(e)}")
            raise
    
    def _create_users_pdf_content(self, data: Dict[str, Any], styles) -> List:
        """Create PDF content for users data."""
        content = []
        
        # Summary
        content.append(Paragraph("User Summary", styles['Heading2']))
        content.append(Paragraph(f"Total Users: {data.get('total_count', 0):,}", styles['Normal']))
        content.append(Spacer(1, 20))
        
        # Users table (first 50 users)
        users = data.get('users', [])[:50]
        if users:
            content.append(Paragraph("User Details (First 50)", styles['Heading2']))
            
            table_data = [['Email', 'Display Name', 'Created At', 'Login Count', 'Total Reports']]
            for user in users:
                table_data.append([
                    user.get('email', '')[:30],  # Truncate long emails
                    user.get('display_name', '')[:20],
                    str(user.get('created_at', ''))[:10],  # Date only
                    str(user.get('login_count', 0)),
                    str(user.get('total_reports_generated', 0))
                ])
            
            table = Table(table_data, colWidths=[2*inch, 1.5*inch, 1*inch, 0.8*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(table)
        
        return content
    
    def _create_audit_logs_pdf_content(self, data: Dict[str, Any], styles) -> List:
        """Create PDF content for audit logs data."""
        content = []
        
        # Summary
        content.append(Paragraph("Audit Log Summary", styles['Heading2']))
        content.append(Paragraph(f"Total Log Entries: {data.get('total_count', 0):,}", styles['Normal']))
        content.append(Spacer(1, 20))
        
        # Recent logs table
        logs = data.get('logs', [])[:20]  # First 20 logs
        if logs:
            content.append(Paragraph("Recent Audit Logs", styles['Heading2']))
            
            table_data = [['Timestamp', 'Admin User', 'Action', 'Target', 'Success']]
            for log in logs:
                table_data.append([
                    str(log.get('timestamp', ''))[:16],  # Date and time
                    log.get('admin_user_id', '')[:15],
                    log.get('action', '')[:20],
                    f"{log.get('target_type', '')}:{log.get('target_id', '')}"[:20],
                    'Yes' if log.get('success') else 'No'
                ])
            
            table = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 1.5*inch, 1.5*inch, 0.6*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(table)
        
        return content
    
    def _create_engagement_pdf_content(self, data: Dict[str, Any], styles, include_charts: bool = False) -> List:
        """Create PDF content for engagement data."""
        content = []
        
        # Overview metrics
        overview = data.get('overview', {})
        if overview:
            content.append(Paragraph("Engagement Overview", styles['Heading2']))
            
            overview_data = [
                ['Metric', 'Value'],
                ['Total Users', f"{overview.get('total_users', 0):,}"],
                ['Active Users', f"{overview.get('active_users', 0):,}"],
                ['Total Sessions', f"{overview.get('total_sessions', 0):,}"],
                ['Avg Session Duration', f"{overview.get('avg_session_duration', 0):.1f} minutes"],
                ['Total Page Views', f"{overview.get('total_page_views', 0):,}"],
                ['Total Reports Generated', f"{overview.get('total_reports_generated', 0):,}"],
                ['Avg Engagement Score', f"{overview.get('avg_engagement_score', 0):.2f}"]
            ]
            
            table = Table(overview_data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(table)
            content.append(Spacer(1, 20))
        
        # Charts (if requested and matplotlib available)
        if include_charts and MATPLOTLIB_AVAILABLE:
            chart_images = self._generate_engagement_charts(data)
            for chart_image in chart_images:
                content.append(chart_image)
                content.append(Spacer(1, 20))
        
        return content
    
    def _create_generic_pdf_content(self, data: Dict[str, Any], styles) -> List:
        """Create generic PDF content for any data type."""
        content = []
        
        content.append(Paragraph("Data Export", styles['Heading2']))
        content.append(Paragraph(f"Data Type: {data.get('data_type', 'Unknown')}", styles['Normal']))
        content.append(Paragraph(f"Generated: {data.get('export_metadata', {}).get('generated_at', '')}", styles['Normal']))
        content.append(Spacer(1, 20))
        
        # Add raw data as formatted text
        content.append(Paragraph("Raw Data (JSON)", styles['Heading3']))
        json_text = json.dumps(data, indent=2, default=str)[:2000]  # Limit size
        if len(json.dumps(data, default=str)) > 2000:
            json_text += "\n... (truncated)"
        
        content.append(Paragraph(f"<pre>{json_text}</pre>", styles['Code']))
        
        return content
    
    def _generate_engagement_charts(self, data: Dict[str, Any]) -> List:
        """Generate chart images for engagement PDF."""
        if not MATPLOTLIB_AVAILABLE:
            return []
        
        charts = []
        
        try:
            plt.style.use('seaborn-v0_8')
            
            # Daily metrics chart
            daily_metrics = data.get('daily_metrics', [])
            if daily_metrics:
                fig, ax = plt.subplots(figsize=(10, 6))
                
                dates = [datetime.fromisoformat(str(metric['date'])) for metric in daily_metrics]
                active_users = [metric['active_users'] for metric in daily_metrics]
                sessions = [metric['total_sessions'] for metric in daily_metrics]
                
                ax.plot(dates, active_users, label='Active Users', marker='o')
                ax.plot(dates, sessions, label='Total Sessions', marker='s')
                
                ax.set_title('Daily Engagement Metrics')
                ax.set_xlabel('Date')
                ax.set_ylabel('Count')
                ax.legend()
                ax.grid(True, alpha=0.3)
                
                # Format dates
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
                ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))
                plt.xticks(rotation=45)
                
                plt.tight_layout()
                
                # Save to buffer
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
                buffer.seek(0)
                
                # Create ReportLab Image
                chart_image = Image(buffer, width=6*inch, height=3.6*inch)
                charts.append(chart_image)
                
                plt.close()
            
        except Exception as e:
            logger.error(f"Error generating engagement charts: {str(e)}")
        
        return charts
    
    async def _export_excel(
        self,
        data: Dict[str, Any],
        export_request: DataExportRequest,
        export_id: str
    ) -> DataExportResult:
        """Export data as Excel file."""
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV format
            logger.warning("Excel export not available - falling back to CSV")
            csv_result = await self._export_csv(data, export_request, export_id)
            csv_result.format = ExportFormat.EXCEL
            csv_result.filename = csv_result.filename.replace('.csv', '.xlsx')
            csv_result.content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return csv_result
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = export_request.data_type.replace('_', ' ').title()
            
            # Add data based on type
            if export_request.data_type == 'users':
                self._create_users_excel_sheet(ws, data)
            elif export_request.data_type == 'audit_logs':
                self._create_audit_logs_excel_sheet(ws, data)
            elif export_request.data_type == 'engagement':
                self._create_engagement_excel_sheet(ws, data)
            else:
                self._create_generic_excel_sheet(ws, data)
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            filename = f"{export_request.data_type}_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return DataExportResult(
                export_id=export_id,
                data_type=export_request.data_type,
                format=ExportFormat.EXCEL,
                filename=filename,
                content=excel_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                size=len(excel_content),
                generated_at=datetime.utcnow()
            )
            
        except Exception as e:
            logger.error(f"Error creating Excel export: {str(e)}")
            raise
    
    def _create_users_excel_sheet(self, ws, data: Dict[str, Any]):
        """Create Excel sheet for users data."""
        # Headers
        headers = [
            'User ID', 'Email', 'Display Name', 'Created At', 'Last Login',
            'Login Count', 'Total Requests', 'Total Reports', 'Screen Time (min)',
            'Subscription Status', 'Geographic Location', 'Engagement Score'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, user in enumerate(data.get('users', []), 2):
            ws.cell(row=row, column=1, value=user.get('user_id', ''))
            ws.cell(row=row, column=2, value=user.get('email', ''))
            ws.cell(row=row, column=3, value=user.get('display_name', ''))
            ws.cell(row=row, column=4, value=user.get('created_at', ''))
            ws.cell(row=row, column=5, value=user.get('last_login', ''))
            ws.cell(row=row, column=6, value=user.get('login_count', 0))
            ws.cell(row=row, column=7, value=user.get('total_requests', 0))
            ws.cell(row=row, column=8, value=user.get('total_reports_generated', 0))
            ws.cell(row=row, column=9, value=user.get('total_screen_time_minutes', 0))
            ws.cell(row=row, column=10, value=user.get('subscription_status', ''))
            ws.cell(row=row, column=11, value=user.get('geographic_location', ''))
            ws.cell(row=row, column=12, value=user.get('engagement_score', 0))
    
    def _create_audit_logs_excel_sheet(self, ws, data: Dict[str, Any]):
        """Create Excel sheet for audit logs data."""
        # Headers
        headers = [
            'Timestamp', 'Admin User ID', 'Action', 'Target Type',
            'Target ID', 'Success', 'IP Address', 'Details'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, log in enumerate(data.get('logs', []), 2):
            ws.cell(row=row, column=1, value=log.get('timestamp', ''))
            ws.cell(row=row, column=2, value=log.get('admin_user_id', ''))
            ws.cell(row=row, column=3, value=log.get('action', ''))
            ws.cell(row=row, column=4, value=log.get('target_type', ''))
            ws.cell(row=row, column=5, value=log.get('target_id', ''))
            ws.cell(row=row, column=6, value='Yes' if log.get('success') else 'No')
            ws.cell(row=row, column=7, value=log.get('ip_address', ''))
            ws.cell(row=row, column=8, value=json.dumps(log.get('details', {})))
    
    def _create_engagement_excel_sheet(self, ws, data: Dict[str, Any]):
        """Create Excel sheet for engagement data."""
        # Overview section
        ws.cell(row=1, column=1, value="ENGAGEMENT OVERVIEW").font = Font(bold=True, size=14)
        
        overview = data.get('overview', {})
        row = 3
        for key, value in overview.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Daily metrics section
        row += 2
        ws.cell(row=row, column=1, value="DAILY METRICS").font = Font(bold=True, size=14)
        row += 2
        
        # Headers for daily metrics
        headers = [
            'Date', 'Total Users', 'Active Users', 'Total Sessions',
            'Total Page Views', 'Total Clicks', 'Reports Generated',
            'Avg Session Duration', 'Engagement Score'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Daily metrics data
        for metric in data.get('daily_metrics', []):
            row += 1
            ws.cell(row=row, column=1, value=metric.get('date', ''))
            ws.cell(row=row, column=2, value=metric.get('total_users', 0))
            ws.cell(row=row, column=3, value=metric.get('active_users', 0))
            ws.cell(row=row, column=4, value=metric.get('total_sessions', 0))
            ws.cell(row=row, column=5, value=metric.get('total_page_views', 0))
            ws.cell(row=row, column=6, value=metric.get('total_clicks', 0))
            ws.cell(row=row, column=7, value=metric.get('total_reports_generated', 0))
            ws.cell(row=row, column=8, value=metric.get('avg_session_duration', 0))
            ws.cell(row=row, column=9, value=metric.get('avg_engagement_score', 0))
    
    def _create_generic_excel_sheet(self, ws, data: Dict[str, Any]):
        """Create generic Excel sheet for any data type."""
        ws.cell(row=1, column=1, value="DATA EXPORT").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value="Data Type")
        ws.cell(row=2, column=2, value=data.get('data_type', 'Unknown'))
        ws.cell(row=3, column=1, value="Generated At")
        ws.cell(row=3, column=2, value=data.get('export_metadata', {}).get('generated_at', ''))
        
        ws.cell(row=5, column=1, value="RAW DATA (JSON)").font = Font(bold=True)
        ws.cell(row=6, column=1, value=json.dumps(data, indent=2, default=str))
    
    async def _send_export_email(
        self,
        export_result: DataExportResult,
        export_request: DataExportRequest
    ) -> bool:
        """Send export file via email."""
        try:
            subject = f"Admin Dashboard Export - {export_request.data_type.replace('_', ' ').title()}"
            
            body = f"""
            Your admin dashboard export is ready.
            
            Export Details:
            - Data Type: {export_request.data_type.replace('_', ' ').title()}
            - Format: {export_request.format.value.upper()}
            - Export ID: {export_result.export_id}
            - Generated: {export_result.generated_at.strftime('%Y-%m-%d %H:%M:%S')} UTC
            - File Size: {export_result.size:,} bytes
            
            Please find the export file attached to this email.
            
            Best regards,
            MINT Admin Team
            """
            
            # Create attachment
            attachment = {
                'filename': export_result.filename,
                'content': base64.b64encode(export_result.content).decode('utf-8'),
                'content_type': export_result.content_type
            }
            
            # Send email to all recipients
            for recipient in export_request.email_recipients:
                await self.email_service.send_email(
                    to_email=recipient,
                    subject=subject,
                    body=body,
                    attachments=[attachment]
                )
            
            logger.info(f"Export email sent to {len(export_request.email_recipients)} recipients")
            return True
            
        except Exception as e:
            logger.error(f"Error sending export email: {str(e)}")
            return False
    
    async def create_scheduled_report(
        self,
        config: ScheduledReportConfig,
        created_by: str
    ) -> ScheduledReport:
        """
        Create a scheduled report configuration.
        
        Args:
            config: Report configuration
            created_by: User who created the schedule
            
        Returns:
            Created scheduled report record
        """
        try:
            # This would integrate with a job scheduler like Celery
            # For now, we'll create a placeholder record
            
            scheduled_report = ScheduledReport(
                id=str(uuid.uuid4()),
                config=config,
                created_by=created_by,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
                run_count=0
            )
            
            logger.info(f"Created scheduled report: {scheduled_report.id}")
            return scheduled_report
            
        except Exception as e:
            logger.error(f"Error creating scheduled report: {str(e)}")
            raise


# Global export service instance
universal_export_service = UniversalExportService()


async def _generate_user_charts(self, data: Dict[str, Any]) -> List[ReportLabImage]:
        """Generate charts for user data."""
        charts = []
        users = data.get('users', [])
        
        if not users:
            return charts
        
        try:
            # User subscription status pie chart
            fig, ax = plt.subplots(figsize=(8, 6))
            
            # Count subscription types
            subscription_counts = {}
            for user in users:
                status = user.get('subscription_status', 'unknown')
                subscription_counts[status] = subscription_counts.get(status, 0) + 1
            
            labels = list(subscription_counts.keys())
            values = list(subscription_counts.values())
            
            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
            ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
            ax.set_title('User Distribution by Subscription Type')
            
            plt.tight_layout()
            
            # Save to buffer
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
            buffer.seek(0)
            
            # Create ReportLab Image
            chart_image = Image(buffer, width=6*inch, height=4*inch)
            charts.append(chart_image)
            
            plt.close()
            
        except Exception as e:
            logger.error(f"Error generating user charts: {str(e)}")
        
        return charts
    
async def _generate_user_analytics_charts(self, data: Dict[str, Any]) -> List[ReportLabImage]:
        """Generate charts for user analytics data."""
        charts = []
        analytics = data.get('analytics', [])
        
        if not analytics:
            return charts
        
        try:
            # User growth over time
            fig, ax = plt.subplots(figsize=(10, 6))
            
            dates = [datetime.fromisoformat(str(analytic['time_bucket'])) for analytic in analytics]
            new_users = [analytic['new_users'] for analytic in analytics]
            active_users = [analytic['active_users'] for analytic in analytics]
            
            ax.plot(dates, new_users, label='New Users', marker='o', linestyle='-', color='#3498db')
            ax.plot(dates, active_users, label='Active Users', marker='s', linestyle='-', color='#2ecc71')
            
            ax.set_title('User Growth and Activity Over Time')
            ax.set_xlabel('Date')
            ax.set_ylabel('Number of Users')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            # Format dates
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))
            plt.xticks(rotation=45)
            
            plt.tight_layout()
            
            # Save to buffer
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
            buffer.seek(0)
            
            # Create ReportLab Image
            chart_image = Image(buffer, width=6*inch, height=3.6*inch)
            charts.append(chart_image)
            
            plt.close()
            
        except Exception as e:
            logger.error(f"Error generating user analytics charts: {str(e)}")
        
        return charts
    
async def _generate_system_metrics_charts(self, data: Dict[str, Any]) -> List[ReportLabImage]:
        """Generate charts for system metrics data."""
        charts = []
        metrics = data.get('metrics', [])
        
        if not metrics:
            return charts
        
        try:
            # Group metrics by type and name
            metric_groups = {}
            for metric in metrics:
                key = f"{metric.get('metric_type', 'unknown')}_{metric.get('metric_name', 'unknown')}"
                if key not in metric_groups:
                    metric_groups[key] = []
                metric_groups[key].append(metric)
            
            # Create a chart for each metric group (limit to 3 charts)
            for i, (key, group) in enumerate(list(metric_groups.items())[:3]):
                fig, ax = plt.subplots(figsize=(10, 6))
                
                dates = [datetime.fromisoformat(str(metric['time_bucket'])) for metric in group]
                avg_values = [metric['avg_value'] for metric in group]
                min_values = [metric['min_value'] for metric in group]
                max_values = [metric['max_value'] for metric in group]
                
                ax.plot(dates, avg_values, label='Average', marker='o', linestyle='-', color='#3498db')
                ax.fill_between(dates, min_values, max_values, alpha=0.2, color='#3498db', label='Min-Max Range')
                
                metric_type, metric_name = key.split('_', 1)
                ax.set_title(f'{metric_type.title()} - {metric_name.title()} Over Time')
                ax.set_xlabel('Date')
                ax.set_ylabel('Value')
                ax.legend()
                ax.grid(True, alpha=0.3)
                
                # Format dates
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d %H:%M'))
                ax.xaxis.set_major_locator(mdates.HourLocator(interval=max(1, len(dates)//10)))
                plt.xticks(rotation=45)
                
                plt.tight_layout()
                
                # Save to buffer
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
                buffer.seek(0)
                
                # Create ReportLab Image
                chart_image = Image(buffer, width=6*inch, height=3.6*inch)
                charts.append(chart_image)
                
                plt.close()
            
        except Exception as e:
            logger.error(f"Error generating system metrics charts: {str(e)}")
        
        return charts
    
async def _generate_audit_logs_charts(self, data: Dict[str, Any]) -> List[ReportLabImage]:
        """Generate charts for audit logs data."""
        charts = []
        logs = data.get('logs', [])
        
        if not logs:
            return charts
        
        try:
            # Action type distribution
            fig, ax = plt.subplots(figsize=(8, 6))
            
            # Count action types
            action_counts = {}
            for log in logs:
                action = log.get('action', 'unknown')
                action_counts[action] = action_counts.get(action, 0) + 1
            
            # Sort by count
            sorted_actions = sorted(action_counts.items(), key=lambda x: x[1], reverse=True)
            labels = [action for action, _ in sorted_actions[:8]]  # Top 8 actions
            values = [count for _, count in sorted_actions[:8]]
            
            bars = ax.bar(labels, values, color='#3498db')
            
            # Add value labels on top of bars
            for bar in bars:
                height = bar.get_height()
                ax.annotate(f'{height}',
                            xy=(bar.get_x() + bar.get_width() / 2, height),
                            xytext=(0, 3),  # 3 points vertical offset
                            textcoords="offset points",
                            ha='center', va='bottom')
            
            ax.set_title('Distribution of Admin Actions')
            ax.set_xlabel('Action Type')
            ax.set_ylabel('Count')
            plt.xticks(rotation=45, ha='right')
            
            plt.tight_layout()
            
            # Save to buffer
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
            buffer.seek(0)
            
            # Create ReportLab Image
            chart_image = Image(buffer, width=6*inch, height=4*inch)
            charts.append(chart_image)
            
            plt.close()
            
        except Exception as e:
            logger.error(f"Error generating audit logs charts: {str(e)}")
        
        return charts
    
async def _generate_engagement_charts(self, data: Dict[str, Any]) -> List[ReportLabImage]:
        """Generate charts for engagement data."""
        charts = []
        daily_metrics = data.get('daily_metrics', [])
        
        if not daily_metrics:
            return charts
        
        try:
            # User engagement over time
            fig, ax = plt.subplots(figsize=(10, 6))
            
            dates = [datetime.fromisoformat(str(metric['date'])) for metric in daily_metrics]
            active_users = [metric['active_users'] for metric in daily_metrics]
            engagement_scores = [metric['avg_engagement_score'] for metric in daily_metrics]
            
            # Create a twin axis
            ax2 = ax.twinx()
            
            # Plot data
            line1 = ax.plot(dates, active_users, label='Active Users', marker='o', linestyle='-', color='#3498db')
            line2 = ax2.plot(dates, engagement_scores, label='Avg Engagement Score', marker='s', linestyle='-', color='#e74c3c')
            
            # Set labels and title
            ax.set_title('User Engagement Metrics Over Time')
            ax.set_xlabel('Date')
            ax.set_ylabel('Active Users')
            ax2.set_ylabel('Engagement Score')
            
            # Combine legends
            lines = line1 + line2
            labels = [l.get_label() for l in lines]
            ax.legend(lines, labels, loc='upper left')
            
            ax.grid(True, alpha=0.3)
            
            # Format dates
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))
            plt.xticks(rotation=45)
            
            plt.tight_layout()
            
            # Save to buffer
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
            buffer.seek(0)
            
            # Create ReportLab Image
            chart_image = Image(buffer, width=6*inch, height=3.6*inch)
            charts.append(chart_image)
            
            plt.close()
            
        except Exception as e:
            logger.error(f"Error generating engagement charts: {str(e)}")
        
        return charts